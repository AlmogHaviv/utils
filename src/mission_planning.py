import pandas as pd

import calendar

import openpyxl

from openpyxl import load_workbook

from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

import os

import argparse

import traceback


### parsers ###
def setup_arg_parser():
    """
    Set up the argument parser.
    """
    parser = argparse.ArgumentParser(description='parse for type of report')
    parser.add_argument('-f', '--full_report', action='store_true', help='If entered, return the full year report')
    parser.add_argument('-y', '--yearly_report', action='store_true', help='If entered, return the  yearly summery report')
    parser.add_argument('-m', '--monthly_report', action='store_true', help='If entered, return the yearly summery until the current month report')
    parser.add_argument('-c', '--company', type=str, help='If entered, specify the company name for the report')
    return parser


def main():
    try:
        """
        Main logic of the script using parsed arguments.
        """
        # Input file names
        jira_data_filename = 'jira-missions-yearly.xlsx'
        planning_data_filename = 'yearly-company-budget.xlsx'

        parser = setup_arg_parser()

        # Parse the arguments
        args = parser.parse_args()

        df = full_dataframe(jira_data_filename, planning_data_filename)

        # Check which report type was selected and perform corresponding action
        if args.full_report:
            print("Generating the full yearly report...")
            # Output file name for the full yearly report
            output_filename = 'yearly-planning.xlsx'
            full_report(df, output_filename)
            adding_months_to_output(output_filename, jira_data_filename, None)
            
        elif args.yearly_report:
            print("Generating the yearly summert report...")
            selcted_columns = ['Total Planned - Yearly', 'Total Time Spent - Yearly', 'Total Difference - Yearly']
            output_filename = 'yearly-data-planning-summery.xlsx'
            col_needed = extract_columns(df, selcted_columns)
            full_report(col_needed, output_filename)

        elif args.monthly_report:
            print("Generating the up to this month summery report...")
            selcted_columns = ['Total Planned', 'Total Time Spent', 'Total Difference']
            output_filename = 'monthly-data-planning-summery.xlsx'
            col_needed = extract_columns(df, selcted_columns)
            full_report(col_needed, output_filename)

        elif args.company:
            print("Generating a report for company:", args.company)
            output_filename = f'yearly-data-planning-{args.company}.xlsx'
            row_needed = filter_by_company_name(df, args.company)
            full_report(row_needed, output_filename)
            adding_months_to_output(output_filename, jira_data_filename, str(args.company))

        else:
            print("please specify the type of report, use -h to know which types there are")
            getting_to_the_right_dir("reports")
            filename = "exception_report.txt"
            with open(filename, "a") as file:
                file.write("Exception occurred in work_ratio.py:\n")
                file.write("Please specify the type of report, use -h to know which types there are\n")
                file.write("Thank You\n")
    
    except UnicodeDecodeError as e:
        print("An exception occurred:", str(e))
        getting_to_the_right_dir("reports")
        filename = "exception_report.txt"
        with open(filename, "a") as file:
            file.write("Exception occurred in work_ratio.py:\n")
            file.write("You have letters that are not in english in the excel file you have inputed in one of the following columns:\n")
            file.write("Assignee,  Custom field (Budget), Issue key, Sprint, Time Spent (Days), Issue id\n")
            file.write("It is probably on Sprint column, please check that the dates are formated like this: 01/01/2023\n")
            file.write("For the developer this is the exception:\n")
            file.write("Thank You\n")

        print(f"An exception occurred. Details saved in {filename}")


    except Exception as e:
        # Handle other exceptions here
        filename = "exception_report.txt"
        getting_to_the_right_dir("reports")
        with open(filename, "a") as file:
            file.write("Exception occurred in main() function:\n")
            file.write(str(e) + "\n")
            file.write("Stack Trace:\n")
            file.write(traceback.format_exc() + "\n")
        
        print(f"An exception occurred. Details saved in {filename}")


def getting_to_the_right_dir(dir_name):
    # gives the path of demo.py
    path = os.path.realpath(__file__)
    # gives the directory where the data exists
    dir = os.path.dirname(path)
  
     #replaces folder names 
    dir = dir.replace('src', dir_name)
  
    # changes the current directory to data 
    os.chdir(dir)


def filter_by_company_name(df, company_name):
    try:
        # Strip any leading or trailing spaces from the provided company name
        company_name = company_name.strip()
        
        # Create a list of valid company names
        valid_names = [
            "AgPlenus",
            "AgSeed",
            "BMB",
            "Biomica",
            "CPB",
            "Canonic",
            "Castera",
            "Chempass",
            "Crispril",
            "LavieBio",
            "MicroBoost",
            "Upkeep",
        ]
        
        # Check if the provided company_name is valid
        if company_name not in valid_names:
            raise ValueError("Please write the name of the company as follows:\n" + "\n".join(valid_names))

        # Filter the DataFrame by the specified company_name
        filtered_df = df[df['Company Name'] == company_name]
        
        return filtered_df
    except Exception as e:
        print("An error occurred:", str(e))
        return None


def extract_columns(df, list_of_columns):
    # Assuming df is your DataFrame
    selected_columns = df.iloc[:, :2]  # Extract the first two columns
    selected_columns = selected_columns.join(df[list_of_columns])
    return selected_columns


def full_report(df, output_filename):
    # Export the full data to an Excel file with merged 'Company Name' cells
    export_dataframe_to_excel_with_merged(df, output_filename)
    
    # Styling the Excel file
    change_excel_font(output_filename, font_name='David', font_size=12, italic=False, color='000000')
    style_excel_file(output_filename)
    color_excel_difference(output_filename)


def full_dataframe(jira_data_filename, planning_data_filename):
    # Name of the 'Company Name' column in your data
    company_column_name = 'Company Name'
    column_of_planned_name = 'Monthly Planned'
    
    # Read the Excel file into a DataFrame
    df = read_excel_file(jira_data_filename)
    
    # Group the data by months
    grouped_data_by_months = divide_by_months(df)
    
    # Initialize a list to store monthly data
    yearly_data = []
    
    # Get the number of months and iterate through them
    last_month = grouped_data_by_months[1] + 1
    for num in range(1, last_month):
        # Get the monthly data and append it to the yearly_data list
        df2 = actual_effort_utilization_monthly(grouped_data_by_months[0], num)
        df1 = transform_yearly_to_monthly_and_divide_by_12(planning_data_filename)
        yearly_data.append(merged_planned_and_actual(df1, df2, [column_of_planned_name]))
    
    # Combine the monthly data into a full yearly DataFrame
    full_data = fully_time_spent_dataframe(yearly_data)
    
    return full_data


# creates the datafram from the jira excel, also creating the company name column using the creating_company_column func and the team column using creating_team_column
def read_excel_file(filename):
    try:
        # gives the path of demo.py
        path = os.path.realpath(__file__)
  
        # gives the directory where the data exists
        dir = os.path.dirname(path)
  
        #replaces folder names 
        dir = dir.replace('src', 'data')
  
        # changes the current directory to data 
        os.chdir(dir)

        # Attempt to read the Excel file into a DataFrame
        df = pd.read_excel(filename)

        # Convert 'Time Spent' from seconds to days
        df['Time Spent (Days)'] = df['Time Spent'] / 3600 / 8  # Convert seconds to days (8-hour workdays)

        # Extract the desired columns by their names
        # Assuming that the Excel file contains columns named 'Time Spent,' 'Sprint,' and 'Assignee'
        time_spent = df['Time Spent (Days)']    
        sprint = df['Sprint']         
        assignee = df['Assignee']
        budget  = df['Custom field (Budget)']
        company_name = creating_company_column(budget)
        team_name = creating_team_column(assignee)
        issue_key = df['Issue key']
        issue_id = df['Issue id']

        # Return the extracted data
        new_df = pd.DataFrame({
            'Time Spent (Days)': time_spent,
            'Sprint': sprint,
            'Assignee': assignee,
            'Custom field (Budget)': budget,
            'Company Name' : company_name,
            'Team Name' : team_name,
            'Issue id' : issue_id,
            'Issue_key' : issue_key,
        })
        return new_df

    except Exception as e:
        # Return None if there's an error while reading the file
        return None
    

# adds to the dataframe the monthly column
def divide_by_months(df):
    # Convert the 'Date' column to datetime
    df['Date'] = pd.to_datetime(df['Sprint'])
    # Create a new column 'Month' to store the month from the 'Date' column
    df['Month'] = df['Date'].dt.month
    # Assign a custom month value to records from December 2022
    df.loc[((df['Month'] == 12) & (df['Date'].dt.year == 2022)) | ((df['Month'] == 11) & (df['Date'].dt.year == 2022)), 'Month'] = 1  # Change November & December 2022 to January 2023
    # Group the data by 'Month'
    grouped_data = dict(tuple(df.groupby('Month')))    # Perform aggregation or analysis on the grouped data, for example, calculate the mean
     # Find the number of the last month
    last_month = df['Month'].max()
    return grouped_data, last_month


# divides the dataframe to dev, algo and bi
def dividing_by_team(grouped_data, month_number):
    if month_number in grouped_data:
        # Filter data for the specified month
        month_data = grouped_data[month_number]
        
        # Create a dictionary to store DataFrames for each group
        group_data = {'Bi': None, 'Algo': None, 'Dev': None}
        
        # Define custom grouping based on 'Assignee'
        custom_groups = {
            'markb': 'Bi',
            'iliab': 'Bi',
            'michala': 'Bi',
            'liorr': 'Bi',
            'robertoo': 'Algo',
            'itair': 'Algo',
            'renanam': 'Algo',
            'anatm': 'Algo',
            'duduz': 'Dev',
            'nerias': 'Dev',
            'noama': 'Dev',
            'hodayam': 'Dev',
        }
        
        # Iterate through unique 'Assignee' values and group accordingly
        for assignee, group_name in custom_groups.items():
            assignee_data = month_data[month_data['Assignee'] == assignee]
            if group_data[group_name] is None:
                group_data[group_name] = assignee_data
            else:
                group_data[group_name] = pd.concat([group_data[group_name], assignee_data])
        
        return group_data
    else:
        return None


# the func using in read_excel_file to creates the coulmn of team name
def creating_team_column(assignee_column):
    # gives the path of demo.py
    path = os.path.realpath(__file__)
  
    # gives the directory where the data exists
    dir = os.path.dirname(path)
    
    #replaces folder names 
    dir = dir.replace('src', 'config')
  
    # changes the current directory to data 
    os.chdir(dir)

    # creating company column
    df = pd.read_csv("worker-names.csv")

    # Extract the desired columns by their names
    dev_names = df["Dev"]
    algo_names = df["Algo"]
    bi_names = df["Bi"]
    devops_name = df["Devops"]

    # creating a list
    res = []


    for name in assignee_column:
        # Remove leading and trailing whitespace from the name
        name = name.strip()
        if name in list(dev_names):
            res.append("Dev") 
        elif name in list(algo_names):
            res.append("Algo")
        elif name in list(bi_names):
            res.append("Bi")
        elif name in list(devops_name):
            res.append("Devops")
        else:
            res.append("Irrelevant")
    return res


# the func using in read_excel_file to creates the company of team name
def creating_company_column(budget_column):
    # gives the path of demo.py
    path = os.path.realpath(__file__)
  
    # gives the directory where the data exists
    dir = os.path.dirname(path)
    
    #replaces folder names 
    dir = dir.replace('src', 'config')
  
    # changes the current directory to data 
    os.chdir(dir)
    
    # creating company column
    df = pd.read_excel("budget-naming.xlsx")

    # Extract the desired columns by their names
    budget_names = df["budget"]
    company_names = df["company name"]

    # creating a dict
    company_codes_dict = {}

    for index, name in enumerate(budget_names):
        company_codes_dict[name] = company_names[index]

    # Desired new list for dataframe
    res = []

    # conditioning
    for item in budget_column:
        # Split the string by '-' and strip any whitespace
        parts = item.split('-')
        code = parts[0].strip()

        # Check if the code exists in the comparison list
        if code in company_codes_dict:
            res.append(company_codes_dict[code])
        else:
            # If not found, you can handle this case as needed
            res.append(-1)  # Or another suitable value

    return res


def actual_effort_utilization_monthly(grouped_data, month_number):
    # Extract the DataFrame for the specified month from the grouped data
    df = grouped_data[month_number]
    
    # Group the DataFrame by 'Company Name' and 'Team Name', and calculate the sum of 'Time Spent (Days)'
    result = df.groupby(['Company Name', 'Team Name'])['Time Spent (Days)'].sum().reset_index()
    
    # Pivot the DataFrame with 'Company Name' as the index, 'Team Name' as columns, and 'Time Spent (Days)' as values
    result_pivot = result.pivot(index='Company Name', columns='Team Name', values='Time Spent (Days)').fillna(0)
    
    # Reset the index to bring 'Company Name' back as a column
    result_pivot.reset_index(inplace=True)
    
    # Melt the DataFrame to reshape it for the desired output
    result_melted = pd.melt(result_pivot, id_vars=['Company Name'], var_name='Team Name', value_name='Time Spent (Days)')
    
    # Sort the DataFrame by 'Company Name' and 'Team Name'
    result_melted.sort_values(by=['Company Name', 'Team Name'], inplace=True)
    
    # Reset the index
    result_melted.reset_index(drop=True, inplace=True)

    month_name = calendar.month_name[month_number]

    result_melted.rename(columns={'Time Spent (Days)': f'{month_name} Time Spent'}, inplace=True)
    
    return (result_melted, f'{month_name} Time Spent')


# Creates he final dataframe for the  excel sheet
def merge_dataframes(df1, df2, column_names1, column_names2):
    # Merge the two DataFrames using an outer join on the specified key columns
    merged_df = pd.merge(df1, df2, on=['Company Name', 'Team Name'], how='outer')

    # Fill NaN values with 0 for the specified columns
    for col in column_names1:
        merged_df[col].fillna(0, inplace=True)
    for col in column_names2:
        merged_df[col].fillna(0, inplace=True)

    return merged_df


# Function to create a full dataframe of all the months of time spent
def fully_time_spent_dataframe(list_of_months_data):
    # Initialize the full dataframe with the data from the first month
    full_data = list_of_months_data[0]

    # Initialize a list to store the names of the months
    column_names = list_of_months_data[0].columns
    column_names_list = column_names.tolist()
    exclude_items = [column_names_list[0], column_names_list[1]]
    list_of_month_names = [item for item in column_names_list if item not in exclude_items]

    # Iterate through the remaining months' data
    for data in list_of_months_data:
        # Check if the month is not January (to avoid merging with itself)
        monthly_column_names = data.columns
        monthly_column_names_list = monthly_column_names.tolist()
        monthly_list_of_names = [item for item in monthly_column_names_list if item not in exclude_items]
        if "January" not in monthly_list_of_names[0]:
            # Separate the full data and the current month's data
            df1 = full_data
            df2 = data

            # Merge the dataframes and update the full data
            full_data = merge_dataframes(df1, df2, list_of_month_names, monthly_list_of_names)

            # Append the current month's name to the list
            list_of_month_names.extend(monthly_list_of_names)

    # Return the full dataframe containing data for all months
    filtered_df = full_data[full_data['Team Name'] != 'Irrelevant']

    # Calculate the sum of planned columns for each row
    planned_columns = filtered_df.columns[filtered_df.columns.str.contains(' Planned')]
    filtered_df['Total Planned'] = filtered_df[planned_columns].sum(axis=1)

    # Calculate the sum of time spent columns for each row
    time_spent_columns = filtered_df.columns[filtered_df.columns.str.contains(' Time Spent')]
    filtered_df['Total Time Spent'] = filtered_df[time_spent_columns].sum(axis=1)

    # Calculate the sum of difference columns for each row
    difference_columns = filtered_df.columns[filtered_df.columns.str.contains(' Difference')]
    filtered_df['Total Difference'] = filtered_df[difference_columns].sum(axis=1)
    

    # Create the new columns for yearly
    filtered_df['Total Planned - Yearly'] = filtered_df['January Planned'] * 12
    filtered_df['Total Time Spent - Yearly'] = filtered_df['Total Time Spent']
    filtered_df['Total Difference - Yearly'] = filtered_df['Total Planned - Yearly'] - filtered_df['Total Time Spent - Yearly']

    # Round up all the numbers
    numeric_columns = filtered_df.select_dtypes(include=[float])
    filtered_df[numeric_columns.columns] = numeric_columns.round(2)

    return filtered_df


# extract time planed for a given month and company from the yearly excel
def transform_yearly_to_monthly_and_divide_by_12(filename):
    # gives the path of demo.py
    path = os.path.realpath(__file__)
  
    # gives the directory where the data exists
    dir = os.path.dirname(path)
    
    #replaces folder names 
    dir = dir.replace('src', 'config')
  
    # changes the current directory to data 
    os.chdir(dir)

    df = pd.read_excel(filename)

    # Melt the DataFrame to convert it to long format
    df = pd.melt(df, id_vars=['Company Name'], var_name='Team Name', value_name='Planned')

    # Divide the planned values by 12 to get monthly planned values
    df['Planned'] /= 12

    # Pivot the DataFrame to have columns for each team
    df = df.pivot(index='Company Name', columns='Team Name', values='Planned').reset_index()

    # Rename the columns for clarity
    df.columns.name = None

    # Fill NaN values with 0
    df.fillna(0, inplace=True)

     # Melt the DataFrame to transform columns to rows
    df = pd.melt(df, id_vars=['Company Name'], var_name='Team Name', value_name='Monthly Planned')

    # Sort the DataFrame by 'Company Name' for better readability
    df.sort_values(by=['Company Name', 'Team Name'], inplace=True)

    # Reset the index
    df.reset_index(drop=True, inplace=True)

    return df


def merged_planned_and_actual(df1, df2, column_names1):
    column_names2 = [df2[1]]
    df2 = df2[0]
    merged_dataframe = merge_dataframes(df1, df2, column_names1, column_names2)
    # Split the string by whitespace
    words = column_names2[0].split()
    # Get the first word
    month = words[0]
    merged_dataframe.rename(columns={'Monthly Planned': f'{month} Planned'}, inplace=True)
    
    # Create the 'Difference' column by subtracting 'month Time Spent' from 'month Planned' 
    merged_dataframe[f'{month} Difference'] = merged_dataframe[f'{month} Planned'] - merged_dataframe[column_names2[0]]

    return  merged_dataframe


def export_dataframe_to_excel_with_merged(df, output_filename):
    # gives the path of the current script
    path = os.path.realpath(__file__)
  
    # gives the directory where the data exists
    dir = os.path.dirname(path)
  
    #replaces folder names 
    dir = dir.replace('src', 'reports')
  
    # changes the current directory to data 
    os.chdir(dir)

    # Get the current working directory as the main folder
    main_folder = os.getcwd()

    # Define the name of the subfolder you want to enter
    subfolder_name = "yearly-report-excel"

    # Create the path to the subfolder
    subfolder_path = os.path.join(main_folder, subfolder_name)

    # Check if the subfolder exists; if not, create it
    if not os.path.exists(subfolder_path):
        os.makedirs(subfolder_path)

    # Now you can save files or outputs to the subfolder
    output_file_path = os.path.join(subfolder_path, output_filename)

    # Create a new Excel writer object
    writer = pd.ExcelWriter(output_file_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Access the Excel workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Iterate through the 'Company Name' column and merge cells
    previous_company = None
    merge_start_row = 2  # Start from row 2
    for row, company in enumerate(df['Company Name'], start=2):
        if company != previous_company:
            if previous_company is not None:
                worksheet.merge_cells(f'A{merge_start_row}:A{row-1}')
            merge_start_row = row
        previous_company = company

    # Merge the last group
    if previous_company is not None:
        worksheet.merge_cells(f'A{merge_start_row}:A{row}')

    # Save the Excel file
    writer.close()

    print(f'Excel file "{output_filename}" created with merged cells.')


def style_excel_file(filename):
    try:
        # Get the current working directory as the main folder
        main_folder = os.getcwd()

        # Define the name of the subfolder you want to enter
        subfolder_name = "yearly-report-excel"

        # Create the path to the subfolder
        subfolder_path = os.path.join(main_folder, subfolder_name)

        # Now you can save files or outputs to the subfolder
        output_file_path = os.path.join(subfolder_path, filename)

        # Load the Excel file using openpyxl
        workbook = load_workbook(output_file_path)

        # Select the active sheet (replace 'Sheet1' with the actual sheet name if different)
        sheet = workbook.active

        # Create a Format object for centering text
        centered_text_format = Alignment(horizontal='center', vertical='center')

        # Apply center alignment to all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = centered_text_format

        # Create a Format object for bold font
        bold_font_format = Font(bold=True)

        # Apply bold font to the 'Company Name' column
        for cell in sheet['A']:
            cell.font = bold_font_format

        # Create a Format object for a bold border
        bold_border_format = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Apply bold border to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = bold_border_format

        # Save the modified Excel file
        workbook.save(output_file_path)
        
        print(f'Styling applied successfully to {filename}')
    except Exception as e:
        print(f'Error: {e}')


def color_excel_difference(filename):
    # Get the current working directory as the main folder
    main_folder = os.getcwd()

    # Define the name of the subfolder you want to enter
    subfolder_name = "yearly-report-excel"

    # Create the path to the subfolder
    subfolder_path = os.path.join(main_folder, subfolder_name)

    # Now you can save files or outputs to the subfolder
    output_file_path = os.path.join(subfolder_path, filename)

    # Load the Excel file
    excel_file = pd.ExcelFile(output_file_path)
    
    # Read the DataFrame from the Excel file
    df = excel_file.parse(excel_file.sheet_names[0])

    # Get the xlsxwriter workbook and worksheet objects.
    workbook = load_workbook(output_file_path)
    worksheet = workbook.active

    # Define the cell fill colors
    positive_color = PatternFill(start_color='D8FFD8', end_color='D8FFD8', fill_type='solid')  # Soft green
    negative_color = PatternFill(start_color='FFD8D8', end_color='FFD8D8', fill_type='solid')  # Soft red
    planned_color = PatternFill(start_color='D8E4FF', end_color='D8E4FF', fill_type='solid')  # Soft blue


    # Loop through the columns and apply cell styles
    for col in df.columns:
        if 'Difference' in col:
            for idx, value in enumerate(df[col], start=2):  # Start from the second row (header is in the first row)
                cell = worksheet.cell(row=idx, column=df.columns.get_loc(col) + 1)  # +1 because Excel is 1-based
                if value > 0:
                    cell.fill = positive_color
                elif value < 0:
                    cell.fill = negative_color
        elif 'Planned' in col:
            for idx, value in enumerate(df[col], start=2):  # Start from the second row (header is in the first row)
                cell = worksheet.cell(row=idx, column=df.columns.get_loc(col) + 1)  # +1 because Excel is 1-based
                cell.fill = planned_color

    # Save the modified Excel file
    workbook.save(output_file_path)


def change_excel_font(filename, font_name='David', font_size=12, italic=False, color='000000'):
    """
    Change the font for the entire Excel file.

    Args:
        filename (str): The name of the Excel file.
        font_name (str, optional): The font name. Default is 'Arial'.
        font_size (int, optional): The font size. Default is 12.
        bold (bool, optional): Whether the font should be bold. Default is False.
        italic (bool, optional): Whether the font should be italic. Default is False.
        color (str, optional): The font color in RGB format. Default is '000000' (black).

    Returns:
        None
    """

    # Get the current working directory as the main folder
    main_folder = os.getcwd()

    # Define the name of the subfolder you want to enter
    subfolder_name = "yearly-report-excel"

    # Create the path to the subfolder
    subfolder_path = os.path.join(main_folder, subfolder_name)

    # Now you can save files or outputs to the subfolder
    output_file_path = os.path.join(subfolder_path, filename)
    # Load the existing Excel workbook
    workbook = load_workbook(output_file_path)

    # Define font settings
    font = Font(name=font_name, size=font_size, italic=italic, color=color)

    # Iterate through all worksheets in the workbook
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]

        # Iterate through all cells in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font

    # Save the modified workbookoutput_file_path
    workbook.save(output_file_path)


def adding_months_to_output(excel_output_filename, jira_filename, company_name):
    getting_to_the_right_dir("src")
    # Read the Excel file into a DataFrame
    df = read_excel_file(jira_filename)
        
    # Group the data by months
    a = divide_by_months(df)
    month_num = a[1]
    grouped_data_by_months = a[0]

    getting_to_the_right_dir("reports")

    # Construct the path to the subfolder
    subfolder_path = os.path.join(os.getcwd(), "yearly-report-excel")

    # Change the current working directory to the subfolder
    os.chdir(subfolder_path)

    # Open the existing Excel file and load the workbook
    book = load_workbook(excel_output_filename)

    # Create a Pandas Excel writer object
    writer = pd.ExcelWriter(excel_output_filename, engine='openpyxl', mode='a')

    for month in grouped_data_by_months:
        grouped_data_by_months[month] = grouped_data_by_months[month][grouped_data_by_months[month]['Team Name'] != 'Irrelevant']
        grouped_data_by_months[month] = grouped_data_by_months[month][['Custom field (Budget)', 'Issue id', 'Issue_key', 'Company Name', 'Team Name', 'Assignee', 'Time Spent (Days)', 'Sprint', 'Date', 'Month']]
        df = grouped_data_by_months[month]
        month_name = calendar.month_name[month]

        if company_name != None:
            filtered_df = df[df['Company Name'].str.contains(company_name)]
            filtered_df = filtered_df.sort_values(['Custom field (Budget)', 'Team Name'], ascending = [True, True])
            # Add the DataFrame to a new sheet in the Excel file
            filtered_df.to_excel(writer, sheet_name=f'{month_name}', index=False)
        else:
            # Add the DataFrame to a new sheet in the Excel file
            df = df.sort_values(['Custom field (Budget)', 'Team Name'], ascending = [True, True])
            df.to_excel(writer, sheet_name=f'{month_name}', index=False)
            print(df)

    writer.close()

    style_sub_sheets(excel_output_filename, month_num)



def style_sub_sheets(excel_output_filename, month_num):
    getting_to_the_right_dir("reports")
    # Construct the path to the subfolder
    subfolder_path = os.path.join(os.getcwd(), "yearly-report-excel")

    # Change the current working directory to the subfolder
    os.chdir(subfolder_path)

    wb = openpyxl.load_workbook(excel_output_filename)

    # Define the border style (thick)
    thick_border = Border(
            top=Side(style='thick', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
            )
    thin_border_format = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

    
    # Define the fill color for the header row (e.g., light gray)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    centered_text_format = Alignment(horizontal='center', vertical='center')

    # Iterate through the rows and add borders between different budget sections
    for num in range(1, month_num+1):
        month_name = calendar.month_name[num]
        sheet = wb[f'{month_name}']

        # Apply bold border to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border_format
                cell.alignment = centered_text_format
        # Iterate through the rows and add borders between different budget sections
        previous_budget = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
            current_budget = row[0].value
            if current_budget != previous_budget:
                for cell in row:
                    cell.border = thick_border
                previous_budget = current_budget
        
        # Apply thin borders to the entire header row (row 1)
        for cell in sheet[1]:
            cell.border = thin_border_format
            cell.fill = header_fill
        

    # Define the border style (thick)
    thick_border = Border(top=Side(style='thick'))

            # Save the modified workbook
    wb.save(excel_output_filename)


### excecution###
if __name__ == "__main__":
    main()
    
'''
# Access the desired sheet by name
    for num in range(month_num):
        month_name = calendar.month_name[num]
        sheet = wb[f'{month_name}']
        # Iterate through the rows and add borders between different budget sections
        previous_budget = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
            current_budget = row[0].value
            if current_budget != previous_budget:
                for cell in row:
                    cell.border = thick_border
                previous_budget = current_budget
'''