import os

import pandas as pd

import mission_planning

import calendar

import openpyxl

from openpyxl import load_workbook

from openpyxl.styles import Alignment, Border, Side, Font, PatternFill



def main():
    # Input and output file names
    jira_data_filename = 'jira-missions-yearly.xlsx'
    planning_budget_filename = 'CPBC.xlsx'
    output_filename = 'budget-and-time-spent-report.xlsx'
    
    # Process Jira data and store it in a dictionary
    monthly_dict = process_jira_data(jira_data_filename)
    
    # Create a 'planned' DataFrame from the budget data
    planned = create_planned_column(planning_budget_filename)
    
    # Reformat the names and merge the data
    integrated_dict_by_month = merged_and_yearly_dict(monthly_dict, planned)

    sum_dict_by_month = adding_sum(integrated_dict_by_month)

    save_dict_of_dataframes_to_excel(sum_dict_by_month, output_filename)

    style_sub_sheets(output_filename)

    # for key, value in integrated_dict_by_month.items():
        # print('------------------------')
        # print(f"sheet name: {key}")
        # print(value)
        # print(planned)



# Function to navigate to the right directory
def getting_to_the_right_dir(dir_name):
    # Get the path of the current script
    path = os.path.realpath(__file__)
    # Extract the directory where the data exists
    dir = os.path.dirname(path)
    # Replace folder names to change the current directory
    dir = dir.replace('src', dir_name)  # Adjust 'src' to the actual folder structure
    # Change the current directory to the specified folder
    os.chdir(dir)
    
    
# Function to create and process a DataFrame
def process_jira_data(data_file):
    getting_to_the_right_dir('data')

    # Read the data from the given Excel file
    df = mission_planning.read_excel_file(data_file)

    # Modify 'Custom field (Budget)' column to remove parentheses and trailing spaces
    df['Custom field (Budget)'] = df['Custom field (Budget)'].str.replace(r'\([^)]*\)', '', regex=True).str.strip()

    # Group the data by months
    month_dataframes = mission_planning.divide_by_months(df)[0]

    # Initialize an empty dictionary to store the result
    team_monthly_data = {}

    # Iterate through each month's DataFrame
    for month, df in month_dataframes.items():
        # Group the DataFrame by the "Team Name" column
        grouped = df.groupby('Team Name')

        # Create a dictionary to store DataFrames for each team within the month
        team_data = {}

        # Iterate through the groups (teams) within the month
        for team, team_df in grouped:
            # Store each team's DataFrame with the team name as the key
            team_data[team] = team_df

        # Store the team-specific DataFrames for this month in the result dictionary
        team_monthly_data[month] = team_data

    # Create an empty dictionary to store the summed DataFrames
    summed_data = {}

    # Iterate through the team_monthly_data dictionary
    for month, team_data in team_monthly_data.items():
        for team, team_df in team_data.items():
            # Group the mini DataFrame by "Custom field (Budget)" and sum "Time Spent (Days)"
            budget_sums = team_df.groupby(['Custom field (Budget)', 'Company Name'])['Time Spent (Days)'].sum()

            # Convert the result to a DataFrame with appropriate column names
            budget_sums_df = budget_sums.reset_index(name='Total Time Spent')

            # Add a 'Team Name' column with the current team name
            budget_sums_df['Team Name'] = team

            # Sort the DataFrame by 'Company Name'
            budget_sums_df.sort_values(by='Company Name', inplace=True)

            # Store the DataFrame in the 'summed_data' dictionary
            if month not in summed_data:
                summed_data[month] = budget_sums_df
            else:
                # Merge the current DataFrame with the existing one for the same month
                summed_data[month] = pd.concat([summed_data[month], budget_sums_df])

    # Return the final 'summed_data' dictionary
    return summed_data


# Reaching the config in order to take all the data about the planned part
def create_planned_column(file_path):

    getting_to_the_right_dir('config')

    # Create an empty DataFrame to store the combined data
    combined_df = pd.DataFrame()

    # Read the Excel file sheet by sheet
    xls = pd.ExcelFile(file_path)
    
    # Loop through each sheet in the Excel file
    for sheet_name in xls.sheet_names:
        # Read the current sheet into a DataFrame
        df = pd.read_excel(xls, sheet_name=sheet_name)
        # Define the starting row (assuming headers are in the first row)
        start_row = 1
        
        # Find the row where "Task Units" and "Project Code" are located
        for index, row in df.iterrows():
            if row[4] == "Task Units" and row[20] == "Project Code":
                start_row = index + 1
                break
        
        
        # Select columns by index (E and U) and rows starting from 'start_row'
        df = df.iloc[start_row:, [4, 8, 20]]
        # Set the first row as column names
        df.columns = df.iloc[0]

        # Remove the first row (since it's now the column names)
        df = df.iloc[1:]

        # Remove rows where "Project Code" is NaN
        df = df.dropna(subset=["Project Code"])
        
        # Replace NaN values with zero
        df.fillna(0, inplace=True)
        
        df["Company Name"] = sheet_name[4:]

        for index, row in df.iterrows():
            if row['Project Code'] == 'P999':
                 df.at[index, 'Project Code'] = f'P999 - {sheet_name[4:]}'
                


        # Append the current sheet's data to the combined DataFrame
        combined_df = pd.concat([df, combined_df], ignore_index=True)
    
    combined_df = combined_df[combined_df['Department'] != 'Product Management']
    combined_df = combined_df[combined_df['Department'] != 'CPB Directors']

    combined_df['Department'].replace({
    'Bioinformatics': 'Bi',
    'Algorithm': 'Algo',
    'Software Development': 'Dev'
    }, inplace=True)

    combined_df.rename(columns={'Project Code': 'Budget', 'Department': 'Team Name', 'Task Units': 'Time planned'}, inplace=True)

    combined_df['Time planned'] = combined_df['Time planned'] / 12

    return combined_df


# Dealing with the budget name P997-
def reformat_string(input_string):
    parts = input_string.split('-')
    if len(parts) >= 2:
        return f'P{parts[-2].strip()} - {"".join(parts[-1:]).strip()}'
    return input_string


# Creatung monthly dict of all the needed data with two yearly dfs - one unitl the last month and one summed yearly plan
def merged_and_yearly_dict(dict_of_dfs, planned):
    new_dict = {}
    for key, df in dict_of_dfs.items():
        for index, row in df.iterrows():
          budget = row['Custom field (Budget)']

          # Check if the budget column starts with 'P997'
          if budget.startswith('P997'):
            
            # Reformat the budget column using the reformat_string function
            df.at[index, 'Custom field (Budget)'] = reformat_string(budget)
        
        # Extract the 'Budget' column from 'Custom field (Budget)'
        df['Budget'] = df['Custom field (Budget)'].str.extract(r'(\w+)')[0]

        # Merge the dataframes based on 'Budget' and 'Team Name'
        merged_df = pd.merge(planned, df, on=['Budget', 'Team Name'], how='outer')
        
        # Fill NaN values in 'Total Time Spent' with 0
        merged_df['Total Time Spent'].fillna(0, inplace=True)
        
        # Load the Excel data into a DataFrame
        getting_to_the_right_dir('config')
        excel_data = pd.read_excel('budget-naming.xlsx')

        # Create a mapping dictionary from 'Budget' to 'Company Name'
        mapping_dict = dict(zip(excel_data['budget'], excel_data['company name']))

        # Replace 'Budget' values in your DataFrame with 'Company Name'
        merged_df['Company Name'] = merged_df['Budget'].map(mapping_dict)

        # Drop the 'Custom field (Budget)' column
        # merged_df = merged_df.drop(['Custom field (Budget)'], axis=1)
        
        # Sort the dataframe
        merged_df.sort_values(by=['Company Name', 'Team Name'], inplace=True)

        # Group by 'Budget' and 'Team Name', then sum the columns
        summed_df = merged_df.groupby(['Budget', 'Team Name'], as_index=False).agg({
            'Company Name': 'first',
            'Time planned': 'sum',
            'Total Time Spent': 'sum'
        })

        # Rearrange the columns
        summed_df = summed_df[['Budget', 'Company Name', 'Team Name', 'Time planned', 'Total Time Spent']]

        # Calculate 'Delta'
        summed_df['Delta'] = summed_df['Time planned'] - summed_df['Total Time Spent']

        new_dict[key] = summed_df 
    
    # Create a list of DataFrames from the values in the dictionary
    dfs_list = list(new_dict.values())

   # Concatenate the DataFrames along the rows (axis=0)
    combined_df = pd.concat(dfs_list, axis=0)

    # Group by 'Budget' and 'Team Name' and sum the other columns
    sum_monthly_df = combined_df.groupby(['Budget', 'Team Name', 'Company Name']).sum().reset_index()

    # Sort the dataframe
    sum_monthly_df.sort_values(by=['Company Name', 'Team Name'], inplace=True)

    # Calculate 'Delta'
    sum_monthly_df['Delta'] = sum_monthly_df['Time planned'] - sum_monthly_df['Total Time Spent']


    # Add the sum DataFrame 
    new_dict['Total - by Month'] = sum_monthly_df

    sum_yearly_df = sum_monthly_df.copy()
    sum_yearly_df['Time planned'] = new_dict[1]['Time planned'] * 12

    # Calculate 'Delta'
    sum_yearly_df['Delta'] = sum_yearly_df['Time planned'] - sum_yearly_df['Total Time Spent']

     # Add the sum DataFrame 
    new_dict['Total - by Year'] = sum_yearly_df


    return new_dict


def adding_sum(dict_of_dfs):
    new_dict = {}
    for key, df in dict_of_dfs.items():
        # Group by 'Company Name' and then by 'Team Name'
        grouped = df.groupby(['Company Name', 'Team Name'])
        # Calculate the sum for each group
        summed_data = grouped[['Time planned', 'Total Time Spent', 'Delta']].sum()
        df2 = pd.DataFrame(summed_data)
        df2['Budget'] = 'A'

        # Merge the summed DataFrame with the original DataFrame
        merged_df = pd.concat([df, df2.reset_index()], ignore_index=True)

        # Sort the merged DataFrame by 'Company Name,' 'Team Name,' and 'Budget'
        sorted_df = merged_df.sort_values(by=['Company Name', 'Team Name', 'Budget'], ascending=[True, True, True])

        if type(key) == int:
            new_key = calendar.month_name[key]
        else:
            new_key = key

        new_dict[new_key] = sorted_df
        

    return new_dict
    

def save_dict_of_dataframes_to_excel(dict_of_dfs, excel_file_name):
    """
    Save a dictionary of DataFrames to an Excel file.

    Parameters:
    - dict_of_dfs: Dictionary where keys are sheet names and values are DataFrames.
    - excel_file_name: Name of the Excel file (e.g., 'output.xlsx').

    Example usage:
    save_dict_of_dataframes_to_excel({'Sheet1': df1, 'Sheet2': df2}, 'output.xlsx')
    """
    # Define the new order of columns
    new_order = ['Company Name', 'Budget', 'Team Name', 'Time planned', 'Total Time Spent', 'Delta']
    for key, df in dict_of_dfs.items():
        # Reorder the columns in the DataFrame
        df = df[new_order]
    getting_to_the_right_dir("reports")

    sheet_order = ['Total - by Year', 'Total - by Month']

    writer = pd.ExcelWriter(excel_file_name, engine='xlsxwriter')

    for sheet_name in sheet_order:
        if sheet_name in dict_of_dfs:
            dict_of_dfs[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
    
    for sheet_name, df in dict_of_dfs.items():
        if sheet_name in dict_of_dfs:
            pass
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    

    writer.close()


def style_sub_sheets(excel_output_filename):
    getting_to_the_right_dir("reports")
    # Construct the path to the subfolder

    wb = openpyxl.load_workbook(excel_output_filename)

    # Define the border style (thick)
    thick_border = Border(
            top=Side(style='thick', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
            )
    thin_border_format = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

    
    # Define the fill color for the header row (e.g., light gray)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    centered_text_format = Alignment(horizontal='center', vertical='center')

    # Get a list of all sheet names in the workbook
    sheet_names = wb.sheetnames

    # Iterate through all the sheets
    for sheet_name in sheet_names:
        sheet = wb[f'{sheet_name}']

        # Apply bold border to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border_format
                cell.alignment = centered_text_format

        # Iterate through the rows and add borders between different Team sections
        previous_team = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
            current_team = row[0].value
            if current_team != previous_team:
                for cell in row:
                    cell.border = thick_border
                previous_team = current_team
        
        # Apply thin borders to the entire header row (row 1)
        for cell in sheet[1]:
            cell.border = thin_border_format
            cell.fill = header_fill
        

    # Define the border style (thick)
    thick_border = Border(top=Side(style='thick'))

            # Save the modified workbook
    wb.save(excel_output_filename)
    print("Sub sheets styls applied")


# Call the main function when the script is executed
if __name__ == '__main__':
    main()
