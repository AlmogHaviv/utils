import pandas as pd

import calendar

from openpyxl import load_workbook

from openpyxl.styles import Alignment, Border, Side, Font, PatternFill


def main():
    # Input and output file names
    jira_data_filename = 'jira-missions-yearly2023.xlsx'
    planning_data_filename = 'yearly-company-budget.xlsx'
    output_filename = 'yearly-planning.xlsx'
    
    # Name of the 'Company Name' column in your data
    company_column_name = 'Company Name'
    column_of_planned_name = 'Monthly Planned'
    
    # Read the Excel file into a DataFrame
    df = read_excel_file(jira_data_filename)
    
    # Group the data by months
    grouped_data_by_months = divide_by_months(df)
    
    # Initialize a list to store monthly data
    yearly_data = []
    
    # Get the number of months and iterate through them
    last_month = grouped_data_by_months[1] + 1
    for num in range(1, last_month):
        # Get the monthly data and append it to the yearly_data list
        df2 = actual_effort_utilization_monthly(grouped_data_by_months[0], num)
        df1 = transform_yearly_to_monthly_and_divide_by_12(planning_data_filename)
        yearly_data.append(merged_planned_and_actual(df1, df2, [column_of_planned_name]))
    
    # Combine the monthly data into a full yearly DataFrame
    full_data = fully_time_spent_dataframe(yearly_data)
    
    print (full_data)

    # Export the full data to an Excel file with merged 'Company Name' cells
    export_dataframe_to_excel_with_merged(full_data, output_filename)

    # Styling the excel file
    change_excel_font(output_filename, font_name='David', font_size=12, italic=False, color='000000')
    style_excel_file(output_filename)
    color_excel_difference(output_filename) 


# creates the datafram from the jira excel, also creating the company name column using the creating_company_column func and the team column using creating_team_column
def read_excel_file(filename):
    try:
        # Attempt to read the Excel file into a DataFrame
        df = pd.read_excel(filename)

        # Convert 'Time Spent' from seconds to days
        df['Time Spent (Days)'] = df['Time Spent'] / 3600 / 8  # Convert seconds to days (8-hour workdays)

        # Extract the desired columns by their names
        # Assuming that the Excel file contains columns named 'Time Spent,' 'Sprint,' and 'Assignee'
        time_spent = df['Time Spent (Days)']    
        sprint = df['Sprint']         
        assignee = df['Assignee']
        budget  = df['Custom field (Budget)']
        company_name = creating_company_column(budget)
        team_name = creating_team_column(assignee)

        # Return the extracted data
        new_df = pd.DataFrame({
            'Time Spent (Days)': time_spent,
            'Sprint': sprint,
            'Assignee': assignee,
            'Custom field (Budget)': budget,
            'Company Name' : company_name,
            'Team Name' : team_name,
        })
        return new_df

    except Exception as e:
        # Return None if there's an error while reading the file
        return None
    

# adds to the dataframe the monthly column
def divide_by_months(df):
    # Convert the 'Date' column to datetime
    df['Date'] = pd.to_datetime(df['Sprint'])
    # Create a new column 'Month' to store the month from the 'Date' column
    df['Month'] = df['Date'].dt.month
    # Assign a custom month value to records from December 2022
    df.loc[((df['Month'] == 12) & (df['Date'].dt.year == 2022)) | ((df['Month'] == 11) & (df['Date'].dt.year == 2022)), 'Month'] = 1  # Change November & December 2022 to January 2023
    # Group the data by 'Month'
    grouped_data = dict(tuple(df.groupby('Month')))    # Perform aggregation or analysis on the grouped data, for example, calculate the mean
     # Find the number of the last month
    last_month = df['Month'].max()
    return grouped_data, last_month


# divides the dataframe to dev, algo and bi
def dividing_by_team(grouped_data, month_number):
    if month_number in grouped_data:
        # Filter data for the specified month
        month_data = grouped_data[month_number]
        
        # Create a dictionary to store DataFrames for each group
        group_data = {'Bi': None, 'Algo': None, 'Dev': None}
        
        # Define custom grouping based on 'Assignee'
        custom_groups = {
            'markb': 'Bi',
            'iliab': 'Bi',
            'michala': 'Bi',
            'liorr': 'Bi',
            'robertoo': 'Algo',
            'itair': 'Algo',
            'renanam': 'Algo',
            'anatm': 'Algo',
            'duduz': 'Dev',
            'nerias': 'Dev',
            'noama': 'Dev',
            'hodayam': 'Dev',
        }
        
        # Iterate through unique 'Assignee' values and group accordingly
        for assignee, group_name in custom_groups.items():
            assignee_data = month_data[month_data['Assignee'] == assignee]
            if group_data[group_name] is None:
                group_data[group_name] = assignee_data
            else:
                group_data[group_name] = pd.concat([group_data[group_name], assignee_data])
        
        return group_data
    else:
        return None


# the func using in read_excel_file to creates the coulmn of team name
def creating_team_column(assignee_column):
    # creating company column
    df = pd.read_csv("worker-names.csv")

    # Extract the desired columns by their names
    dev_names = df["Dev"]
    algo_names = df["Algo"]
    bi_names = df["Bi"]
    devops_name = df["Devops"]

    # creating a list
    res = []


    for name in assignee_column:
        # Remove leading and trailing whitespace from the name
        name = name.strip()
        if name in list(dev_names):
            res.append("Dev") 
        elif name in list(algo_names):
            res.append("Algo")
        elif name in list(bi_names):
            res.append("Bi")
        elif name in list(devops_name):
            res.append("Devops")
        else:
            res.append("Irrelevant")
    return res


# the func using in read_excel_file to creates the company of team name
def creating_company_column(budget_column):
    # creating company column
    df = pd.read_excel("budget-naming.xlsx")

    # Extract the desired columns by their names
    budget_names = df["budget"]
    company_names = df["company name"]

    # creating a dict
    company_codes_dict = {}

    for index, name in enumerate(budget_names):
        company_codes_dict[name] = company_names[index]

    # Desired new list for dataframe
    res = []

    # conditioning
    for item in budget_column:
        # Split the string by '-' and strip any whitespace
        parts = item.split('-')
        code = parts[0].strip()

        # Check if the code exists in the comparison list
        if code in company_codes_dict:
            res.append(company_codes_dict[code])
        else:
            # If not found, you can handle this case as needed
            res.append(-1)  # Or another suitable value

    return res


def actual_effort_utilization_monthly(grouped_data, month_number):
    # Extract the DataFrame for the specified month from the grouped data
    df = grouped_data[month_number]
    
    # Group the DataFrame by 'Company Name' and 'Team Name', and calculate the sum of 'Time Spent (Days)'
    result = df.groupby(['Company Name', 'Team Name'])['Time Spent (Days)'].sum().reset_index()
    
    # Pivot the DataFrame with 'Company Name' as the index, 'Team Name' as columns, and 'Time Spent (Days)' as values
    result_pivot = result.pivot(index='Company Name', columns='Team Name', values='Time Spent (Days)').fillna(0)
    
    # Reset the index to bring 'Company Name' back as a column
    result_pivot.reset_index(inplace=True)
    
    # Melt the DataFrame to reshape it for the desired output
    result_melted = pd.melt(result_pivot, id_vars=['Company Name'], var_name='Team Name', value_name='Time Spent (Days)')
    
    # Sort the DataFrame by 'Company Name' and 'Team Name'
    result_melted.sort_values(by=['Company Name', 'Team Name'], inplace=True)
    
    # Reset the index
    result_melted.reset_index(drop=True, inplace=True)

    month_name = calendar.month_name[month_number]

    result_melted.rename(columns={'Time Spent (Days)': f'{month_name} Time Spent'}, inplace=True)
    
    return (result_melted, f'{month_name} Time Spent')


# Creates he final dataframe for the  excel sheet
def merge_dataframes(df1, df2, column_names1, column_names2):
    # Merge the two DataFrames using an outer join on the specified key columns
    merged_df = pd.merge(df1, df2, on=['Company Name', 'Team Name'], how='outer')

    # Fill NaN values with 0 for the specified columns
    for col in column_names1:
        merged_df[col].fillna(0, inplace=True)
    for col in column_names2:
        merged_df[col].fillna(0, inplace=True)

    return merged_df


# Function to create a full dataframe of all the months of time spent
def fully_time_spent_dataframe(list_of_months_data):
    # Initialize the full dataframe with the data from the first month
    full_data = list_of_months_data[0]

    # Initialize a list to store the names of the months
    column_names = list_of_months_data[0].columns
    column_names_list = column_names.tolist()
    exclude_items = [column_names_list[0], column_names_list[1]]
    list_of_month_names = [item for item in column_names_list if item not in exclude_items]

    # Iterate through the remaining months' data
    for data in list_of_months_data:
        # Check if the month is not January (to avoid merging with itself)
        monthly_column_names = data.columns
        monthly_column_names_list = monthly_column_names.tolist()
        monthly_list_of_names = [item for item in monthly_column_names_list if item not in exclude_items]
        if "January" not in monthly_list_of_names[0]:
            # Separate the full data and the current month's data
            df1 = full_data
            df2 = data

            # Merge the dataframes and update the full data
            full_data = merge_dataframes(df1, df2, list_of_month_names, monthly_list_of_names)

            # Append the current month's name to the list
            list_of_month_names.extend(monthly_list_of_names)

    # Return the full dataframe containing data for all months
    filtered_df = full_data[full_data['Team Name'] != 'Irrelevant']

    # Calculate the sum of planned columns for each row
    planned_columns = filtered_df.columns[filtered_df.columns.str.contains(' Planned')]
    filtered_df['Total Planned'] = filtered_df[planned_columns].sum(axis=1)

    # Calculate the sum of time spent columns for each row
    time_spent_columns = filtered_df.columns[filtered_df.columns.str.contains(' Time Spent')]
    filtered_df['Total Time Spent'] = filtered_df[time_spent_columns].sum(axis=1)

    # Calculate the sum of difference columns for each row
    difference_columns = filtered_df.columns[filtered_df.columns.str.contains(' Difference')]
    filtered_df['Total Difference'] = filtered_df[difference_columns].sum(axis=1)
    

    # Create the new columns for yearly
    filtered_df['Total Planned - Yearly'] = filtered_df['January Planned'] * 12
    filtered_df['Total Time Spent - Yearly'] = filtered_df['Total Time Spent']
    filtered_df['Total Difference - Yearly'] = filtered_df['Total Planned - Yearly'] - filtered_df['Total Time Spent - Yearly']

    # Round up all the numbers
    numeric_columns = filtered_df.select_dtypes(include=[float])
    filtered_df[numeric_columns.columns] = numeric_columns.round(2)

    return filtered_df


# extract time planed for a given month and company from the yearly excel
def transform_yearly_to_monthly_and_divide_by_12(filename):
    df = pd.read_excel(filename)

    # Melt the DataFrame to convert it to long format
    df = pd.melt(df, id_vars=['Company Name'], var_name='Team Name', value_name='Planned')

    # Divide the planned values by 12 to get monthly planned values
    df['Planned'] /= 12

    # Pivot the DataFrame to have columns for each team
    df = df.pivot(index='Company Name', columns='Team Name', values='Planned').reset_index()

    # Rename the columns for clarity
    df.columns.name = None

    # Fill NaN values with 0
    df.fillna(0, inplace=True)

     # Melt the DataFrame to transform columns to rows
    df = pd.melt(df, id_vars=['Company Name'], var_name='Team Name', value_name='Monthly Planned')

    # Sort the DataFrame by 'Company Name' for better readability
    df.sort_values(by=['Company Name', 'Team Name'], inplace=True)

    # Reset the index
    df.reset_index(drop=True, inplace=True)

    return df


def merged_planned_and_actual(df1, df2, column_names1):
    column_names2 = [df2[1]]
    df2 = df2[0]
    merged_dataframe = merge_dataframes(df1, df2, column_names1, column_names2)
    # Split the string by whitespace
    words = column_names2[0].split()
    # Get the first word
    month = words[0]
    merged_dataframe.rename(columns={'Monthly Planned': f'{month} Planned'}, inplace=True)
    
    # Create the 'Difference' column by subtracting 'month Time Spent' from 'month Planned' 
    merged_dataframe[f'{month} Difference'] = merged_dataframe[f'{month} Planned'] - merged_dataframe[column_names2[0]]

    return  merged_dataframe


def export_dataframe_to_excel_with_merged(df, output_filename):
    # Create a new Excel writer object
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Access the Excel workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Iterate through the 'Company Name' column and merge cells
    previous_company = None
    merge_start_row = 2  # Start from row 2
    for row, company in enumerate(df['Company Name'], start=2):
        if company != previous_company:
            if previous_company is not None:
                worksheet.merge_cells(f'A{merge_start_row}:A{row-1}')
            merge_start_row = row
        previous_company = company

    # Merge the last group
    if previous_company is not None:
        worksheet.merge_cells(f'A{merge_start_row}:A{row}')

    # Save the Excel file
    writer.close()

    print(f'Excel file "{output_filename}" created with merged cells.')


def style_excel_file(filename):
    try:
        # Load the Excel file using openpyxl
        workbook = load_workbook(filename)

        # Select the active sheet (replace 'Sheet1' with the actual sheet name if different)
        sheet = workbook.active

        # Create a Format object for centering text
        centered_text_format = Alignment(horizontal='center', vertical='center')

        # Apply center alignment to all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = centered_text_format

        # Create a Format object for bold font
        bold_font_format = Font(bold=True)

        # Apply bold font to the 'Company Name' column
        for cell in sheet['A']:
            cell.font = bold_font_format

        # Create a Format object for a bold border
        bold_border_format = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Apply bold border to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = bold_border_format

        # Save the modified Excel file
        workbook.save(filename)
        
        print(f'Styling applied successfully to {filename}')
    except Exception as e:
        print(f'Error: {e}')


def color_excel_difference(filename):
    # Load the Excel file
    excel_file = pd.ExcelFile(filename)
    
    # Read the DataFrame from the Excel file
    df = excel_file.parse(excel_file.sheet_names[0])

    # Get the xlsxwriter workbook and worksheet objects.
    workbook = load_workbook(filename)
    worksheet = workbook.active

    # Define the cell fill colors
    positive_color = PatternFill(start_color='D8FFD8', end_color='D8FFD8', fill_type='solid')  # Soft green
    negative_color = PatternFill(start_color='FFD8D8', end_color='FFD8D8', fill_type='solid')  # Soft red
    planned_color = PatternFill(start_color='D8E4FF', end_color='D8E4FF', fill_type='solid')  # Soft blue


    # Loop through the columns and apply cell styles
    for col in df.columns:
        if 'Difference' in col:
            for idx, value in enumerate(df[col], start=2):  # Start from the second row (header is in the first row)
                cell = worksheet.cell(row=idx, column=df.columns.get_loc(col) + 1)  # +1 because Excel is 1-based
                if value > 0:
                    cell.fill = positive_color
                elif value < 0:
                    cell.fill = negative_color
        elif 'Planned' in col:
            for idx, value in enumerate(df[col], start=2):  # Start from the second row (header is in the first row)
                cell = worksheet.cell(row=idx, column=df.columns.get_loc(col) + 1)  # +1 because Excel is 1-based
                cell.fill = planned_color

    # Save the modified Excel file
    workbook.save(filename)


def change_excel_font(filename, font_name='David', font_size=12, italic=False, color='000000'):
    """
    Change the font for the entire Excel file.

    Args:
        filename (str): The name of the Excel file.
        font_name (str, optional): The font name. Default is 'Arial'.
        font_size (int, optional): The font size. Default is 12.
        bold (bool, optional): Whether the font should be bold. Default is False.
        italic (bool, optional): Whether the font should be italic. Default is False.
        color (str, optional): The font color in RGB format. Default is '000000' (black).

    Returns:
        None
    """
    # Load the existing Excel workbook
    workbook = load_workbook(filename)

    # Define font settings
    font = Font(name=font_name, size=font_size, italic=italic, color=color)

    # Iterate through all worksheets in the workbook
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]

        # Iterate through all cells in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font

    # Save the modified workbook
    workbook.save(filename)

### excecution###
if __name__ == "__main__":
    main()
    