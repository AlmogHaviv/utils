from openpyxl import load_workbook
from openpyxl.styles import Font

def change_excel_font(filename, font_name='David', font_size=12, italic=False, color='000000'):
    """
    Change the font for the entire Excel file.

    Args:
        filename (str): The name of the Excel file.
        font_name (str, optional): The font name. Default is 'Arial'.
        font_size (int, optional): The font size. Default is 12.
        bold (bool, optional): Whether the font should be bold. Default is False.
        italic (bool, optional): Whether the font should be italic. Default is False.
        color (str, optional): The font color in RGB format. Default is '000000' (black).

    Returns:
        None
    """
    # Load the existing Excel workbook
    workbook = load_workbook(filename)

    # Define font settings
    font = Font(name=font_name, size=font_size, italic=italic, color=color)

    # Iterate through all worksheets in the workbook
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]

        # Iterate through all cells in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font

    # Save the modified workbook
    workbook.save(filename)

# Usage example:
change_excel_font('yearly-planning.xlsx', font_name='Verdana', color='000000')
